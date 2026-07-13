from __future__ import annotations

import csv
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from scytaledroid.DynamicAnalysis.ml.publication_bundle.exporters import (
    tex_escape,
    write_csv_with_provenance,
    write_tex_table,
    write_xlsx,
)
from scytaledroid.DynamicAnalysis.ml.artifact_bundle_writer import _paper_provenance


def test_write_csv_with_provenance_comment_header(tmp_path: Path) -> None:
    path = tmp_path / "table.csv"

    write_csv_with_provenance(
        path,
        ["package_name", "score"],
        [{"package_name": "com.example_app", "score": 1.5}],
        provenance={"freeze_sha256": "abc123"},
    )

    text = path.read_text(encoding="utf-8")
    assert text.startswith("# freeze_sha256: abc123\n#\n")
    rows = list(csv.DictReader(line for line in text.splitlines() if not line.startswith("#")))
    assert rows == [{"package_name": "com.example_app", "score": "1.5"}]


def test_write_tex_table_escapes_headers_and_values(tmp_path: Path) -> None:
    path = tmp_path / "table.tex"

    write_tex_table(
        path,
        columns=[("package_name", "Package_Name"), ("note", "Note")],
        rows=[{"package_name": "com.example_app", "note": r"path\value"}],
        provenance={"source": "unit_test"},
        caption_comment="Caption_text",
    )

    text = path.read_text(encoding="utf-8")
    assert "% source: unit_test" in text
    assert "Package\\_Name" in text
    assert "com.example\\_app" in text
    assert "path\\textbackslash{}value" in text
    assert tex_escape("a_b") == "a\\_b"


def test_write_xlsx_has_provenance_and_table_sheets(tmp_path: Path) -> None:
    path = tmp_path / "table.xlsx"

    write_xlsx(
        path,
        sheet_name="table_1",
        columns=[("package_name", "Package"), ("score", "Score")],
        rows=[{"package_name": "com.example", "score": 2}],
        provenance={"freeze_sha256": "abc123"},
    )

    wb = load_workbook(path)
    assert wb.sheetnames == ["provenance", "table_1"]
    assert wb["provenance"]["A2"].value == "freeze_sha256"
    assert wb["provenance"]["B2"].value == "abc123"
    assert wb["table_1"]["A1"].value == "Package"
    assert wb["table_1"]["B2"].value == 2


def test_write_xlsx_is_byte_stable_for_same_content(tmp_path: Path) -> None:
    kwargs = {
        "sheet_name": "table_1",
        "columns": [("package_name", "Package"), ("score", "Score")],
        "rows": [{"package_name": "com.example", "score": 2}],
        "provenance": {"freeze_sha256": "abc123"},
    }
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    write_xlsx(first, **kwargs)
    write_xlsx(second, **kwargs)

    assert first.read_bytes() == second.read_bytes()
    with ZipFile(first) as workbook:
        assert {info.date_time for info in workbook.infolist()} == {(2000, 1, 1, 0, 0, 0)}
        assert b"<dcterms:modified" in workbook.read("docProps/core.xml")
        assert b">2000-01-01T00:00:00Z</dcterms:modified>" in workbook.read("docProps/core.xml")


def test_paper_table_provenance_excludes_generation_timestamp() -> None:
    provenance = _paper_provenance(freeze_sha256="abc123")

    assert provenance["freeze_sha256"] == "abc123"
    assert "generated_at_utc" not in provenance
