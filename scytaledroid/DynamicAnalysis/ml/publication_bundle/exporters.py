"""Format exporters for runtime ML publication bundle tables."""

from __future__ import annotations

import csv
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def write_csv_with_provenance(
    path: Path, fieldnames: list[str], rows: list[dict[str, Any]], *, provenance: dict[str, str]
) -> None:
    """Write a CSV with a provenance comment header."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        for k, v in provenance.items():
            handle.write(f"# {k}: {v}\n")
        handle.write("#\n")
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k) for k in fieldnames})


def tex_escape(s: str) -> str:
    return s.replace("\\", "\\textbackslash{}").replace("_", "\\_")


def write_tex_table(
    path: Path,
    *,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    provenance: dict[str, str],
    caption_comment: str,
) -> None:
    """Write a standalone LaTeX tabular with provenance comments."""
    path.parent.mkdir(parents=True, exist_ok=True)
    keys = [k for k, _ in columns]
    headers = [tex_escape(h) for _, h in columns]
    spec = "l" + ("r" * (len(columns) - 1))
    with path.open("w", encoding="utf-8") as handle:
        for k, v in provenance.items():
            handle.write(f"% {k}: {v}\n")
        handle.write("%\n")
        handle.write(f"% {caption_comment}\n")
        handle.write(f"\\begin{{tabular}}{{{spec}}}\n")
        handle.write("\\hline\n")
        handle.write(" & ".join(headers) + " \\\\\n")
        handle.write("\\hline\n")
        for row in rows:
            vals: list[str] = []
            for k in keys:
                v = row.get(k)
                if v is None:
                    vals.append("-")
                else:
                    vals.append(tex_escape(str(v)))
            handle.write(" & ".join(vals) + " \\\\\n")
        handle.write("\\hline\n")
        handle.write("\\end{tabular}\n")


def write_xlsx(
    path: Path,
    *,
    sheet_name: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    provenance: dict[str, str],
) -> None:
    """Write an XLSX with a provenance sheet and one table sheet."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    stable_timestamp = datetime(2000, 1, 1, tzinfo=UTC)
    wb.properties.creator = "ScytaleDroid"
    wb.properties.lastModifiedBy = "ScytaleDroid"
    wb.properties.created = stable_timestamp
    wb.properties.modified = stable_timestamp
    ws0 = wb.active
    ws0.title = "provenance"
    ws0["A1"] = "key"
    ws0["B1"] = "value"
    ws0["A1"].font = Font(bold=True)
    ws0["B1"].font = Font(bold=True)
    i = 2
    for k, v in provenance.items():
        ws0[f"A{i}"] = k
        ws0[f"B{i}"] = v
        i += 1
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 80

    ws = wb.create_sheet(title=sheet_name)
    headers = [h for _, h in columns]
    keys = [k for k, _ in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(k) for k in keys])
    wb.save(path)
    _normalize_xlsx_zip_metadata(path)


def _normalize_xlsx_zip_metadata(path: Path) -> None:
    """Rewrite the XLSX container with stable ZIP timestamps and permissions."""

    fixed_date = (2000, 1, 1, 0, 0, 0)
    with ZipFile(path, "r") as zin:
        entries = [(info, zin.read(info.filename)) for info in zin.infolist()]
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as zout:
        for old_info, data in entries:
            if old_info.filename == "docProps/core.xml":
                data = re.sub(
                    rb"<dcterms:modified([^>]*)>.*?</dcterms:modified>",
                    rb"<dcterms:modified\1>2000-01-01T00:00:00Z</dcterms:modified>",
                    data,
                )
            info = ZipInfo(old_info.filename, fixed_date)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            zout.writestr(info, data)


__all__ = [
    "tex_escape",
    "write_csv_with_provenance",
    "write_tex_table",
    "write_xlsx",
]
