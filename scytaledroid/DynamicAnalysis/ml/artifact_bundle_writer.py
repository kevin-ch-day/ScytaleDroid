"""Generate Phase E (Paper #2) deliverable bundle under output/.

This does NOT mutate evidence packs and does NOT change the freeze anchor.
It packages already-derived tables and generates paper-facing figures/tables,
plus a reproducibility appendix snippet and a manifest with hashes.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

# Force a headless-safe backend. This module may run in CLI/CI environments
# without a display server.
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402

from scytaledroid.Config import app_config

from . import ml_parameters_paper2 as config
from .deliverable_bundle_paths import (
    dataset_tables_dir,
    freeze_anchor_path,
    output_paper_appendix_dir,
    output_paper_artifacts_manifest_path,
    output_paper_figures_dir,
    output_paper_freeze_copy_path,
    output_paper_manifest_dir,
    output_paper_readme_path,
    output_paper_root,
    output_paper_tables_dir,
)
from .pcap_window_features import build_window_features, extract_packet_timeline
from .evidence_pack_ml_preflight import get_sampling_duration_seconds, load_run_inputs
from .telemetry_windowing import WindowSpec


@dataclass(frozen=True)
class PhaseEArtifacts:
    out_root: Path
    fig_b1_png: Path
    fig_b1_pdf: Path
    table_1_csv: Path
    table_2_csv: Path
    table_3_csv: Path
    table_4_csv: Path
    table_5_csv: Path
    table_6_csv: Path
    table_7_csv: Path
    repro_appendix_md: Path
    artifacts_manifest_json: Path
    generated_at: str


def write_phase_e_deliverables_bundle(
    *,
    fig_b1_run_id: str,
    interaction_tag: str | None = None,
) -> PhaseEArtifacts:
    """Write Phase E deliverables under output/paper/paper2/phase_e/.

    Assumes:
    - Freeze anchor exists and is checksummed.
    - Runner already generated canonical dataset CSVs under data/.
    - ML v1 outputs exist for the exemplar run (anomaly_scores_*).
    """

    out_root = output_paper_root()
    figs_dir = output_paper_figures_dir()
    tables_dir = output_paper_tables_dir()
    appendix_dir = output_paper_appendix_dir()
    manifest_dir = output_paper_manifest_dir()
    for d in (out_root, figs_dir, tables_dir, appendix_dir, manifest_dir):
        d.mkdir(parents=True, exist_ok=True)

    # Keep the deliverable bundle clean and "ship-able": remove legacy/obsolete outputs
    # from prior iterations so operators don't accidentally cite the wrong file.
    _clean_bundle_dirs(tables_dir=tables_dir, figs_dir=figs_dir, fig_b1_run_id=fig_b1_run_id)

    freeze_sha256 = _sha256_stream(freeze_anchor_path())
    provenance = _paper_provenance(freeze_sha256=freeze_sha256)

    # Tables (PM locked): emit csv + xlsx + tex under output/paper/...
    table_1_csv, table_1_xlsx, table_1_tex = _write_table_1_rdi_prevalence(tables_dir, provenance=provenance)
    table_2_csv, table_2_xlsx, table_2_tex = _write_table_2_transport_mix(tables_dir, provenance=provenance)
    table_3_csv, table_3_xlsx, table_3_tex = _write_table_3_model_agreement(tables_dir, provenance=provenance)
    table_4_csv, table_4_xlsx, table_4_tex = _write_table_4_signature_deltas(tables_dir, provenance=provenance)
    table_5_csv, table_5_xlsx, table_5_tex, masvs_inputs = _write_table_5_masvs_coverage(
        tables_dir, provenance=provenance
    )
    table_6_csv, table_6_xlsx, table_6_tex = _write_table_6_static_posture_scores(
        tables_dir, provenance=provenance
    )
    table_7_csv, table_7_xlsx, table_7_tex = _write_table_7_exposure_deviation_summary(
        tables_dir, provenance=provenance
    )

    # Freeze anchor copy (convenience; canonical stays in data/archive/).
    _copy_required(freeze_anchor_path(), output_paper_freeze_copy_path(), overwrite=True)

    # Copy the exemplar pin lockfile for audit convenience.
    from .evidence_pack_ml_orchestrator import PAPER_ARTIFACTS_PATH

    _copy_required(PAPER_ARTIFACTS_PATH, output_paper_manifest_dir() / "paper_artifacts.json", overwrite=True)

    # Figures (PM locked): Fig B1 + B2 + B4.
    fig_b1_png, fig_b1_pdf = _write_fig_b1(fig_b1_run_id, figs_dir, interaction_tag=interaction_tag, overwrite=True)
    fig_b2_png, fig_b2_pdf = _write_fig_b2(figs_dir, provenance=provenance, overwrite=True)
    fig_b4_png, fig_b4_pdf = _write_fig_b4(figs_dir, provenance=provenance, overwrite=True)

    # Repro appendix snippet.
    repro_appendix_md = appendix_dir / "repro_appendix_phase_e.md"
    repro_appendix_md.write_text(_render_repro_appendix(), encoding="utf-8")

    # Bundle README.
    readme = output_paper_readme_path()
    readme.write_text(_render_bundle_readme(fig_b1_run_id), encoding="utf-8")

    # Bundle manifest (hashes + versions + pointers).
    artifacts_manifest_json = output_paper_artifacts_manifest_path()
    _write_bundle_manifest(
        artifacts_manifest_json,
        fig_b1_run_id=fig_b1_run_id,
        fig_b1_png=fig_b1_png,
        fig_b1_pdf=fig_b1_pdf,
        table_1_csv=table_1_csv,
        table_1_xlsx=table_1_xlsx,
        table_1_tex=table_1_tex,
        table_2_csv=table_2_csv,
        table_2_xlsx=table_2_xlsx,
        table_2_tex=table_2_tex,
        table_3_csv=table_3_csv,
        table_3_xlsx=table_3_xlsx,
        table_3_tex=table_3_tex,
        table_4_csv=table_4_csv,
        table_4_xlsx=table_4_xlsx,
        table_4_tex=table_4_tex,
        table_5_csv=table_5_csv,
        table_5_xlsx=table_5_xlsx,
        table_5_tex=table_5_tex,
        table_6_csv=table_6_csv,
        table_6_xlsx=table_6_xlsx,
        table_6_tex=table_6_tex,
        table_7_csv=table_7_csv,
        table_7_xlsx=table_7_xlsx,
        table_7_tex=table_7_tex,
        repro_appendix_md=repro_appendix_md,
        fig_b2_png=fig_b2_png,
        fig_b2_pdf=fig_b2_pdf,
        fig_b4_png=fig_b4_png,
        fig_b4_pdf=fig_b4_pdf,
        masvs_inputs=masvs_inputs,
    )

    # Close-out receipt: pins freeze + bundle-manifest hashes so a zipped bundle
    # can be verified later without re-running anything.
    _write_bundle_closure_record(
        output_paper_manifest_dir() / "phase_e_closure_record.json",
        bundle_manifest_path=artifacts_manifest_json,
    )

    return PhaseEArtifacts(
        out_root=out_root,
        fig_b1_png=fig_b1_png,
        fig_b1_pdf=fig_b1_pdf,
        table_1_csv=table_1_csv,
        table_2_csv=table_2_csv,
        table_3_csv=table_3_csv,
        table_4_csv=table_4_csv,
        table_5_csv=table_5_csv,
        table_6_csv=table_6_csv,
        table_7_csv=table_7_csv,
        repro_appendix_md=repro_appendix_md,
        artifacts_manifest_json=artifacts_manifest_json,
        generated_at=datetime.now(UTC).isoformat(),
    )


def _write_fig_b1(
    fig_run_id: str, figs_dir: Path, *, interaction_tag: str | None, overwrite: bool
) -> tuple[Path, Path]:
    stem = f"fig_b1_timeline_{fig_run_id[:8]}"
    png = figs_dir / f"{stem}.png"
    pdf = figs_dir / f"{stem}.pdf"
    if not overwrite and png.exists() and pdf.exists():
        return png, pdf

    run_dir = Path(app_config.OUTPUT_DIR) / "evidence" / "dynamic" / fig_run_id
    inputs = load_run_inputs(run_dir)
    if not inputs:
        raise RuntimeError(f"Fig B1 run missing run_manifest.json: {fig_run_id}")
    if not inputs.pcap_path or not inputs.pcap_path.exists():
        raise RuntimeError(f"Fig B1 run missing PCAP: {fig_run_id}")

    duration_s = get_sampling_duration_seconds(inputs)
    if duration_s is None or duration_s <= 0:
        raise RuntimeError(f"Fig B1 run missing sampling duration: {fig_run_id}")

    spec = WindowSpec(window_size_s=config.WINDOW_SIZE_S, stride_s=config.WINDOW_STRIDE_S)
    packets = extract_packet_timeline(inputs.pcap_path)
    rows, dropped = build_window_features(packets, duration_s=float(duration_s), spec=spec)
    if not rows:
        raise RuntimeError(f"Fig B1 run produced 0 windows: {fig_run_id}")

    denom = float(spec.window_size_s) if spec.window_size_s > 0 else 1.0
    xs = [(float(r["window_start_s"]) + float(r["window_end_s"])) / 2.0 for r in rows]
    bytes_ps = [float(r["byte_count"]) / denom for r in rows]
    pkts_ps = [float(r["packet_count"]) / denom for r in rows]

    # Load anomaly flags from both models (v1 outputs).
    out_dir = run_dir / "analysis" / "ml" / config.ML_SCHEMA_LABEL
    if_csv = out_dir / "anomaly_scores_iforest.csv"
    oc_csv = out_dir / "anomaly_scores_ocsvm.csv"
    if not if_csv.exists() or not oc_csv.exists():
        raise RuntimeError(f"Fig B1 run missing v1 anomaly score CSVs: {fig_run_id}")
    if_flags = _read_flags(if_csv)
    oc_flags = _read_flags(oc_csv)
    if len(if_flags) != len(xs) or len(oc_flags) != len(xs):
        raise RuntimeError(
            "Fig B1 invariant failed: anomaly score rows must match window count "
            f"(windows={len(xs)} if_rows={len(if_flags)} oc_rows={len(oc_flags)}). "
            "Refusing to generate a misleading figure."
        )

    pkg = inputs.package_name or "<unknown>"
    tag = interaction_tag or _interaction_tag(inputs.manifest) or "interactive"
    title = f"Fig B1: {pkg} ({tag})"

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10.5, 6.5), sharex=True)
    fig.suptitle(title, fontsize=12)

    ax1.plot(xs, bytes_ps, color="#0B3C5D", linewidth=1.5)
    ax1.set_ylabel("Bytes/sec")
    ax1.grid(True, alpha=0.25)

    ax2.plot(xs, pkts_ps, color="#328CC1", linewidth=1.5)
    ax2.set_ylabel("Packets/sec")
    ax2.set_xlabel("Time (s, relative to capture start)")
    ax2.grid(True, alpha=0.25)

    def _overlay(ax, ys, flags, marker, color, label):
        y_max = max(ys) if ys else 0.0
        y_plot = y_max * 1.02 if y_max > 0 else 1.0
        xs_f = [x for x, f in zip(xs, flags, strict=True) if f]
        ax.scatter(xs_f, [y_plot for _ in xs_f], s=22, marker=marker, color=color, alpha=0.85, label=label)

    _overlay(ax1, bytes_ps, if_flags, marker="^", color="#D1495B", label="IF (flagged)")
    _overlay(ax1, bytes_ps, oc_flags, marker="o", color="#00798C", label="OC-SVM (flagged)")
    ax1.legend(loc="upper right", fontsize=8, frameon=False)

    note = f"dropped_partial_windows={dropped} window={spec.window_size_s}s stride={spec.stride_s}s"
    ax2.text(0.01, 0.02, note, transform=ax2.transAxes, fontsize=8, alpha=0.8)

    fig.tight_layout(rect=[0, 0.02, 1, 0.95])
    fig.savefig(png, dpi=200)
    fig.savefig(pdf)
    plt.close(fig)
    return png, pdf


def _write_fig_b2(figs_dir: Path, *, provenance: dict[str, str], overwrite: bool) -> tuple[Path, Path]:
    """Fig B2: RDI prevalence by app (idle vs interactive), faceted by model."""
    stem = "fig_b2_rdi_by_app"
    png = figs_dir / f"{stem}.png"
    pdf = figs_dir / f"{stem}.pdf"
    if not overwrite and png.exists() and pdf.exists():
        return png, pdf

    rows = _read_csv_rows(dataset_tables_dir() / "anomaly_prevalence_per_app_phase.csv")
    apps = sorted({r.get("package_name") for r in rows if r.get("package_name")})
    labels = [config.DISPLAY_NAME_BY_PACKAGE.get(a, a) for a in apps]

    def get(pkg: str, phase: str, model: str) -> float:
        for r in rows:
            if r.get("package_name") == pkg and r.get("phase") == phase and r.get("model") == model:
                try:
                    return float(r.get("flagged_pct") or 0.0)
                except Exception:
                    return 0.0
        return 0.0

    x = np.arange(len(apps))
    width = 0.35
    colors = {"idle": "#0B3C5D", "interactive": "#328CC1"}

    fig, axes = plt.subplots(2, 1, figsize=(12.5, 7.2), sharex=True)
    for ax, (model_key, model_label) in zip(
        axes, [(config.MODEL_IFOREST, "Isolation Forest"), (config.MODEL_OCSVM, "OC-SVM")], strict=True
    ):
        idle_vals = [get(a, "idle", model_key) for a in apps]
        int_vals = [get(a, "interactive", model_key) for a in apps]
        ax.bar(x - width / 2, idle_vals, width, label="Idle", color=colors["idle"], alpha=0.92)
        ax.bar(x + width / 2, int_vals, width, label="Interactive", color=colors["interactive"], alpha=0.92)
        ax.set_ylabel("RDI (flagged %)")
        ax.set_title(model_label, fontsize=10)
        ax.grid(True, axis="y", alpha=0.25)
        ax.set_ylim(0.0, 1.0)
    axes[0].legend(loc="upper right", fontsize=9, frameon=False)
    axes[-1].set_xticks(x, labels, rotation=35, ha="right")
    axes[-1].set_xlabel("App")
    fig.suptitle("Fig B2: Runtime Deviation Index (RDI) by App (Idle vs Interactive)", fontsize=12)
    fig.tight_layout(rect=[0, 0.02, 1, 0.95])
    fig.savefig(png, dpi=200)
    fig.savefig(pdf)
    plt.close(fig)
    return png, pdf


def _write_fig_b4(figs_dir: Path, *, provenance: dict[str, str], overwrite: bool) -> tuple[Path, Path]:
    """Fig B4: static posture vs interactive RDI (context only)."""
    stem = "fig_b4_static_vs_rdi"
    png = figs_dir / f"{stem}.png"
    pdf = figs_dir / f"{stem}.pdf"
    if not overwrite and png.exists() and pdf.exists():
        return png, pdf

    posture = _compute_static_posture_scores()
    rdi = _load_interactive_rdi_iforest()
    pkgs = sorted(set(posture.keys()) & set(rdi.keys()))
    xs = [posture[p][0] for p in pkgs]
    ys = [rdi[p] for p in pkgs]

    rho = _spearman_rho(xs, ys)

    fig, ax = plt.subplots(1, 1, figsize=(9.5, 6.2))
    ax.scatter(xs, ys, s=46, color="#0B3C5D", alpha=0.86)
    ax.set_xlabel("Static Posture Score (0–100, context only)")
    ax.set_ylabel("Interactive RDI (IF flagged %)")
    ax.grid(True, alpha=0.25)
    ax.set_ylim(0.0, 1.0)

    if xs and ys:
        ax.axvline(float(np.median(xs)), color="#888888", linewidth=1.0, alpha=0.7)
        ax.axhline(float(np.median(ys)), color="#888888", linewidth=1.0, alpha=0.7)

    for pkg, x0, y0 in zip(pkgs, xs, ys, strict=True):
        name = config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg)
        ax.text(x0 + 0.8, y0 + 0.01, name, fontsize=8, alpha=0.85)

    subtitle = "Spearman rho: n/a" if rho is None else f"Spearman rho={rho:.2f}"
    ax.set_title("Fig B4: Static Posture vs Runtime Deviation (Discordance)", fontsize=12)
    ax.text(0.01, 0.02, subtitle, transform=ax.transAxes, fontsize=9, alpha=0.85)
    fig.tight_layout()
    fig.savefig(png, dpi=200)
    fig.savefig(pdf)
    plt.close(fig)
    return png, pdf


def _render_repro_appendix() -> str:
    # Keep it simple; this is a paper snippet, not a full report.
    freeze = freeze_anchor_path()
    sha = _sha256_stream(freeze)
    return (
        "# Phase E Reproducibility (Generated)\n\n"
        f"- Freeze anchor: `{freeze}`\n"
        f"- Freeze sha256: `{sha}`\n"
        f"- Windowing: `{config.WINDOW_SIZE_S}s` window / `{config.WINDOW_STRIDE_S}s` stride (drop partials)\n"
        f"- MIN_WINDOWS_BASELINE: `{config.MIN_WINDOWS_BASELINE}`\n"
        f"- Thresholding: `{config.THRESHOLD_PERCENTILE}th percentile` per model×app\n"
        "- Score semantics: higher = more anomalous (normalized)\n"
        "- Selection/training/scoring: DB-free; evidence packs + freeze anchor only\n"
    )


def _render_bundle_readme(fig_run_id: str) -> str:
    return (
        "# Paper #2 Phase E Deliverables Bundle\n\n"
        "This folder is operator/paper-facing. It is intended to be zipped and shared.\n\n"
        "Authoritative inputs:\n"
        "- Evidence packs under `output/evidence/dynamic/<run_id>/...`\n"
        f"- Freeze anchor (canonical): `{freeze_anchor_path()}`\n"
        "- Copy of freeze anchor is included under `manifest/dataset_freeze.json` for convenience.\n\n"
        "Contents:\n"
        f"- figures/: Fig B1 timeline for exemplar run `{fig_run_id}`\n"
        "- figures/: Fig B2 prevalence summary; Fig B4 static-vs-dynamic discordance\n"
        "- tables/: Table 1–7 (csv + xlsx + tex)\n"
        "- appendix/: reproducibility snippet\n"
        "- manifest/: bundle manifest with hashes and pointers\n"
        "- manifest/: closure record (`phase_e_closure_record.json`) pins freeze + artifact hashes\n"
    )


def _write_bundle_manifest(
    path: Path,
    *,
    fig_b1_run_id: str,
    fig_b1_png: Path,
    fig_b1_pdf: Path,
    table_1_csv: Path,
    table_1_xlsx: Path,
    table_1_tex: Path,
    table_2_csv: Path,
    table_2_xlsx: Path,
    table_2_tex: Path,
    table_3_csv: Path,
    table_3_xlsx: Path,
    table_3_tex: Path,
    table_4_csv: Path,
    table_4_xlsx: Path,
    table_4_tex: Path,
    table_5_csv: Path,
    table_5_xlsx: Path,
    table_5_tex: Path,
    table_6_csv: Path,
    table_6_xlsx: Path,
    table_6_tex: Path,
    table_7_csv: Path,
    table_7_xlsx: Path,
    table_7_tex: Path,
    repro_appendix_md: Path,
    fig_b2_png: Path,
    fig_b2_pdf: Path,
    fig_b4_png: Path,
    fig_b4_pdf: Path,
    masvs_inputs: list[dict[str, str]] | None = None,
) -> None:
    payload = {
        "generated_at": datetime.now(UTC).isoformat(),
        "freeze_anchor": str(freeze_anchor_path()),
        "freeze_sha256": _sha256_stream(freeze_anchor_path()),
        "fig_b1_run_id": fig_b1_run_id,
        "ml_schema_version": config.ML_SCHEMA_VERSION,
        "report_schema_version": config.REPORT_SCHEMA_VERSION,
        # Table 5 is derived from static reports (Phase B) using the plan's base_apk_sha256.
        # This is context-only; it does not affect Phase E training/scoring.
        "masvs_inputs": masvs_inputs or [],
        "files": {
            "fig_b1_png": {"path": str(fig_b1_png), "sha256": _sha256_stream(fig_b1_png)},
            "fig_b1_pdf": {"path": str(fig_b1_pdf), "sha256": _sha256_stream(fig_b1_pdf)},
            "fig_b2_png": {"path": str(fig_b2_png), "sha256": _sha256_stream(fig_b2_png)},
            "fig_b2_pdf": {"path": str(fig_b2_pdf), "sha256": _sha256_stream(fig_b2_pdf)},
            "fig_b4_png": {"path": str(fig_b4_png), "sha256": _sha256_stream(fig_b4_png)},
            "fig_b4_pdf": {"path": str(fig_b4_pdf), "sha256": _sha256_stream(fig_b4_pdf)},
            "table_1_csv": {"path": str(table_1_csv), "sha256": _sha256_stream(table_1_csv)},
            "table_1_xlsx": {"path": str(table_1_xlsx), "sha256": _sha256_stream(table_1_xlsx)},
            "table_1_tex": {"path": str(table_1_tex), "sha256": _sha256_stream(table_1_tex)},
            "table_2_csv": {"path": str(table_2_csv), "sha256": _sha256_stream(table_2_csv)},
            "table_2_xlsx": {"path": str(table_2_xlsx), "sha256": _sha256_stream(table_2_xlsx)},
            "table_2_tex": {"path": str(table_2_tex), "sha256": _sha256_stream(table_2_tex)},
            "table_3_csv": {"path": str(table_3_csv), "sha256": _sha256_stream(table_3_csv)},
            "table_3_xlsx": {"path": str(table_3_xlsx), "sha256": _sha256_stream(table_3_xlsx)},
            "table_3_tex": {"path": str(table_3_tex), "sha256": _sha256_stream(table_3_tex)},
            "table_4_csv": {"path": str(table_4_csv), "sha256": _sha256_stream(table_4_csv)},
            "table_4_xlsx": {"path": str(table_4_xlsx), "sha256": _sha256_stream(table_4_xlsx)},
            "table_4_tex": {"path": str(table_4_tex), "sha256": _sha256_stream(table_4_tex)},
            "table_5_csv": {"path": str(table_5_csv), "sha256": _sha256_stream(table_5_csv)},
            "table_5_xlsx": {"path": str(table_5_xlsx), "sha256": _sha256_stream(table_5_xlsx)},
            "table_5_tex": {"path": str(table_5_tex), "sha256": _sha256_stream(table_5_tex)},
            "table_6_csv": {"path": str(table_6_csv), "sha256": _sha256_stream(table_6_csv)},
            "table_6_xlsx": {"path": str(table_6_xlsx), "sha256": _sha256_stream(table_6_xlsx)},
            "table_6_tex": {"path": str(table_6_tex), "sha256": _sha256_stream(table_6_tex)},
            "table_7_csv": {"path": str(table_7_csv), "sha256": _sha256_stream(table_7_csv)},
            "table_7_xlsx": {"path": str(table_7_xlsx), "sha256": _sha256_stream(table_7_xlsx)},
            "table_7_tex": {"path": str(table_7_tex), "sha256": _sha256_stream(table_7_tex)},
            "repro_appendix": {"path": str(repro_appendix_md), "sha256": _sha256_stream(repro_appendix_md)},
            "freeze_copy": {
                "path": str(output_paper_freeze_copy_path()),
                "sha256": _sha256_stream(output_paper_freeze_copy_path()),
            },
            "paper_artifacts_copy": {
                "path": str(output_paper_manifest_dir() / "paper_artifacts.json"),
                "sha256": _sha256_stream(output_paper_manifest_dir() / "paper_artifacts.json"),
            },
        },
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _copy_required(src: Path, dest: Path, *, overwrite: bool) -> None:
    if not src.exists():
        raise RuntimeError(f"Missing required input for bundle: {src}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and not overwrite:
        return
    dest.write_bytes(src.read_bytes())


def _write_bundle_closure_record(path: Path, *, bundle_manifest_path: Path) -> None:
    """Write a close-out receipt for the current bundle contents.

    This file exists so a zipped `output/paper/.../phase_e/` bundle can be verified
    later without depending on any DB state or rerunning generation.
    """
    from .evidence_pack_ml_orchestrator import PAPER_ARTIFACTS_PATH
    from scytaledroid.Utils.toolchain_versions import gather_toolchain_versions

    payload = {
        "bundle_manifest_path": str(bundle_manifest_path),
        "bundle_manifest_sha256": _sha256_stream(bundle_manifest_path),
        "bundle_root": str(output_paper_root()),
        "closed_at_utc": datetime.now(UTC).isoformat(),
        "toolchain": gather_toolchain_versions(),
        "freeze_anchor": str(freeze_anchor_path()),
        "freeze_sha256": _sha256_stream(freeze_anchor_path()),
        "ml_schema_version": int(config.ML_SCHEMA_VERSION),
        "paper_artifacts_path": str(PAPER_ARTIFACTS_PATH),
        "paper_artifacts_sha256": _sha256_stream(PAPER_ARTIFACTS_PATH),
        "report_schema_version": int(config.REPORT_SCHEMA_VERSION),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _sha256_stream(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _interaction_tag(manifest: dict[str, Any]) -> str | None:
    op = manifest.get("operator") if isinstance(manifest.get("operator"), dict) else {}
    msg = str(op.get("messaging_activity") or "").strip().lower()
    if msg:
        return msg
    inter = str(op.get("interaction_level") or "").strip().lower()
    return inter or None


def _read_flags(path: Path) -> list[bool]:
    out: list[bool] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            v = str(row.get("is_anomalous") or "").strip().lower()
            out.append(v in ("1", "true", "t", "yes", "y"))
    return out


def _paper_provenance(*, freeze_sha256: str) -> dict[str, str]:
    """Key/value provenance embedded into paper-facing outputs (CSV/TEX headers)."""
    return {
        "freeze_anchor": str(freeze_anchor_path()),
        "freeze_sha256": str(freeze_sha256),
        "ml_schema_version": str(config.ML_SCHEMA_VERSION),
        "report_schema_version": str(config.REPORT_SCHEMA_VERSION),
        "generated_at_utc": datetime.now(UTC).isoformat(),
    }


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [dict(r) for r in reader]


def _read_csv_rows_with_comment_header(path: Path) -> list[dict[str, str]]:
    """Read a CSV that may start with provenance comment lines beginning with '#'.

    Our paper-facing CSVs embed a comment header; `csv.DictReader` does not skip comments.
    """
    text = path.read_text(encoding="utf-8")
    lines = []
    for line in text.splitlines():
        if not lines and (line.startswith("#") or not line.strip()):
            continue
        lines.append(line)
    if not lines:
        return []
    buf = io.StringIO("\n".join(lines) + "\n")
    reader = csv.DictReader(buf)
    return [dict(r) for r in reader]


def _clean_bundle_dirs(*, tables_dir: Path, figs_dir: Path, fig_b1_run_id: str) -> None:
    """Remove legacy files from previous bundle iterations.

    The Phase E bundle is a paper-facing deliverable; leaving old filenames around
    creates a real risk of citing stale artifacts.
    """
    keep_table_stems = {
        "table_1_rdi_prevalence",
        "table_2_transport_mix",
        "table_3_model_agreement",
        "table_4_signature_deltas",
        "table_5_masvs_coverage",
        "table_6_static_posture_scores",
        "table_7_exposure_deviation_summary",
    }
    keep_fig_stems = {
        f"fig_b1_timeline_{fig_b1_run_id[:8]}",
        "fig_b2_rdi_by_app",
        "fig_b4_static_vs_rdi",
    }
    keep_exts = {".csv", ".xlsx", ".tex", ".png", ".pdf"}

    for d, keep in ((tables_dir, keep_table_stems), (figs_dir, keep_fig_stems)):
        for p in d.iterdir():
            if not p.is_file():
                continue
            if p.suffix not in keep_exts:
                continue
            stem = p.stem
            if d == figs_dir and stem.startswith("fig_b1_timeline_"):
                # Only keep the currently pinned exemplar.
                if stem != f"fig_b1_timeline_{fig_b1_run_id[:8]}":
                    try:
                        p.unlink()
                    except Exception:
                        pass
                continue
            if stem.startswith("table_") or stem.startswith("fig_"):
                if stem not in keep:
                    try:
                        p.unlink()
                    except Exception:
                        pass


def _write_csv_with_provenance(
    path: Path, fieldnames: list[str], rows: list[dict[str, Any]], *, provenance: dict[str, str]
) -> None:
    """Write a CSV with a provenance comment header (paper-grade)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        for k, v in provenance.items():
            handle.write(f"# {k}: {v}\n")
        handle.write("#\n")
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k) for k in fieldnames})


def _tex_escape(s: str) -> str:
    return s.replace("\\", "\\textbackslash{}").replace("_", "\\_")


def _write_tex_table(
    path: Path,
    *,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    provenance: dict[str, str],
    caption_comment: str,
) -> None:
    """Write a standalone LaTeX tabular (no preamble, no table env)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    keys = [k for k, _ in columns]
    headers = [_tex_escape(h) for _, h in columns]
    spec = "l" + ("r" * (len(columns) - 1))
    with path.open("w", encoding="utf-8") as handle:
        for k, v in provenance.items():
            handle.write(f"% {k}: {v}\n")
        handle.write("%\n")
        handle.write(f"% {caption_comment}\n")
        handle.write(f"\\begin{{tabular}}{{{spec}}}\n")
        handle.write("\\hline\n")
        handle.write(" & ".join(headers) + " \\\\\n")
        handle.write("\\hline\n")
        for r in rows:
            vals: list[str] = []
            for k in keys:
                v = r.get(k)
                if v is None:
                    vals.append("-")
                else:
                    vals.append(_tex_escape(str(v)))
            handle.write(" & ".join(vals) + " \\\\\n")
        handle.write("\\hline\n")
        handle.write("\\end{tabular}\n")


def _write_xlsx(
    path: Path,
    *,
    sheet_name: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    provenance: dict[str, str],
) -> None:
    """Write an XLSX with a provenance sheet + a table sheet."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "provenance"
    ws0["A1"] = "key"
    ws0["B1"] = "value"
    ws0["A1"].font = Font(bold=True)
    ws0["B1"].font = Font(bold=True)
    i = 2
    for k, v in provenance.items():
        ws0[f"A{i}"] = k
        ws0[f"B{i}"] = v
        i += 1
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 80

    ws = wb.create_sheet(title=sheet_name)
    headers = [h for _, h in columns]
    keys = [k for k, _ in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(k) for k in keys])
    wb.save(path)


def _write_table_1_rdi_prevalence(tables_dir: Path, *, provenance: dict[str, str]) -> tuple[Path, Path, Path]:
    rows = _read_csv_rows(dataset_tables_dir() / "anomaly_prevalence_per_app_phase.csv")
    pkgs = sorted({r.get("package_name") for r in rows if r.get("package_name")})

    def pick(pkg: str, phase: str, model: str) -> dict[str, str] | None:
        for r in rows:
            if r.get("package_name") == pkg and r.get("phase") == phase and r.get("model") == model:
                return r
        return None

    def fnum(v: str | None) -> float:
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    out: list[dict[str, Any]] = []
    for pkg in pkgs:
        idle_if = pick(pkg, "idle", config.MODEL_IFOREST)
        int_if = pick(pkg, "interactive", config.MODEL_IFOREST)
        idle_oc = pick(pkg, "idle", config.MODEL_OCSVM)
        int_oc = pick(pkg, "interactive", config.MODEL_OCSVM)
        if not (idle_if and int_if and idle_oc and int_oc):
            continue
        out.append(
            {
                "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
                "if_idle": round(fnum(idle_if.get("flagged_pct")), 3),
                "if_interactive": round(fnum(int_if.get("flagged_pct")), 3),
                "oc_idle": round(fnum(idle_oc.get("flagged_pct")), 3),
                "oc_interactive": round(fnum(int_oc.get("flagged_pct")), 3),
                "idle_windows": int(fnum(idle_if.get("windows_total"))),
                "interactive_windows": int(fnum(int_if.get("windows_total"))),
                "training_mode": str(int_if.get("training_mode") or "baseline_only"),
            }
        )

    cols = [
        ("app", "App"),
        ("if_idle", "IF Idle"),
        ("if_interactive", "IF Interactive"),
        ("oc_idle", "OCSVM Idle"),
        ("oc_interactive", "OCSVM Interactive"),
        ("idle_windows", "Idle n"),
        ("interactive_windows", "Interactive n"),
        ("training_mode", "Train"),
    ]
    csv_path = tables_dir / "table_1_rdi_prevalence.csv"
    xlsx_path = tables_dir / "table_1_rdi_prevalence.xlsx"
    tex_path = tables_dir / "table_1_rdi_prevalence.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_1", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment="Table 1: RDI prevalence (flagged %) per app (idle vs interactive).",
    )
    return csv_path, xlsx_path, tex_path


def _write_table_2_transport_mix(tables_dir: Path, *, provenance: dict[str, str]) -> tuple[Path, Path, Path]:
    rows = _read_csv_rows(dataset_tables_dir() / "transport_mix_by_phase.csv")
    pkgs = sorted({r.get("package_name") for r in rows if r.get("package_name")})

    def pick(pkg: str, phase: str) -> dict[str, str] | None:
        for r in rows:
            if r.get("package_name") == pkg and r.get("phase") == phase:
                return r
        return None

    out: list[dict[str, Any]] = []
    for pkg in pkgs:
        idle = pick(pkg, "idle")
        inter = pick(pkg, "interactive")
        if not idle or not inter:
            continue

        def ff(v: str | None) -> float | None:
            try:
                return round(float(v or 0.0), 3)
            except Exception:
                return None

        out.append(
            {
                "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
                "idle_tls": ff(idle.get("tls_ratio")),
                "idle_quic": ff(idle.get("quic_ratio")),
                "idle_tcp": ff(idle.get("tcp_ratio")),
                "idle_udp": ff(idle.get("udp_ratio")),
                "int_tls": ff(inter.get("tls_ratio")),
                "int_quic": ff(inter.get("quic_ratio")),
                "int_tcp": ff(inter.get("tcp_ratio")),
                "int_udp": ff(inter.get("udp_ratio")),
            }
        )

    cols = [
        ("app", "App"),
        ("idle_tls", "Idle TLS"),
        ("idle_quic", "Idle QUIC"),
        ("idle_tcp", "Idle TCP"),
        ("idle_udp", "Idle UDP"),
        ("int_tls", "Int TLS"),
        ("int_quic", "Int QUIC"),
        ("int_tcp", "Int TCP"),
        ("int_udp", "Int UDP"),
    ]
    csv_path = tables_dir / "table_2_transport_mix.csv"
    xlsx_path = tables_dir / "table_2_transport_mix.xlsx"
    tex_path = tables_dir / "table_2_transport_mix.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_2", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment="Table 2: Transport mix context (idle vs interactive).",
    )
    return csv_path, xlsx_path, tex_path


def _write_table_3_model_agreement(tables_dir: Path, *, provenance: dict[str, str]) -> tuple[Path, Path, Path]:
    rows = _read_csv_rows(dataset_tables_dir() / "model_overlap_per_run.csv")
    pkgs = sorted({r.get("package_name") for r in rows if r.get("package_name")})

    def as_int(v: str | None) -> int:
        try:
            return int(float(v or 0))
        except Exception:
            return 0

    out: list[dict[str, Any]] = []
    for pkg in pkgs:
        rs = [r for r in rows if r.get("package_name") == pkg and str(r.get("phase") or "").startswith("interactive")]
        if not rs:
            continue
        both = sum(as_int(r.get("both_flagged")) for r in rs)
        either = sum(as_int(r.get("either_flagged")) for r in rs)
        windows = sum(as_int(r.get("windows_total")) for r in rs)
        jaccard = (float(both) / float(either)) if either > 0 else 0.0
        disagree = max(either - both, 0)
        disagree_rate = (float(disagree) / float(either)) if either > 0 else 0.0
        training_mode = str(rs[0].get("training_mode") or "baseline_only")
        out.append(
            {
                "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
                "jaccard": round(jaccard, 3),
                "disagree_rate": round(disagree_rate, 3),
                "either_flagged": int(either),
                "both_flagged": int(both),
                "interactive_windows": int(windows),
                "training_mode": training_mode,
            }
        )

    cols = [
        ("app", "App"),
        ("jaccard", "Jaccard"),
        ("disagree_rate", "Disagree rate"),
        ("either_flagged", "Either n"),
        ("both_flagged", "Both n"),
        ("interactive_windows", "Int windows"),
        ("training_mode", "Train"),
    ]
    csv_path = tables_dir / "table_3_model_agreement.csv"
    xlsx_path = tables_dir / "table_3_model_agreement.xlsx"
    tex_path = tables_dir / "table_3_model_agreement.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_3", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment="Table 3: IF vs OC-SVM agreement on interactive windows (concatenated).",
    )
    return csv_path, xlsx_path, tex_path


def _quantile(xs: list[float], q: float) -> float:
    if not xs:
        return 0.0
    arr = np.array(xs, dtype=float)
    return float(np.quantile(arr, q, method="linear"))


def _spearman_rho(xs: list[float], ys: list[float]) -> float | None:
    """Best-effort Spearman correlation without scipy (small n, ties handled by average rank)."""
    if len(xs) != len(ys) or len(xs) < 2:
        return None

    def ranks(vals: list[float]) -> list[float]:
        order = sorted(range(len(vals)), key=lambda i: (vals[i], i))
        out = [0.0] * len(vals)
        i = 0
        while i < len(order):
            j = i
            v = vals[order[i]]
            while j < len(order) and vals[order[j]] == v:
                j += 1
            # average rank (1-based)
            avg = (i + 1 + j) / 2.0
            for k in range(i, j):
                out[order[k]] = avg
            i = j
        return out

    rx = np.array(ranks(xs), dtype=float)
    ry = np.array(ranks(ys), dtype=float)
    # Pearson correlation on ranks.
    dx = rx - float(rx.mean())
    dy = ry - float(ry.mean())
    denom = float(np.sqrt((dx * dx).sum()) * np.sqrt((dy * dy).sum()))
    if denom <= 0:
        return None
    return float((dx * dy).sum() / denom)


def _write_table_4_signature_deltas(tables_dir: Path, *, provenance: dict[str, str]) -> tuple[Path, Path, Path]:
    """Compute per-app descriptive deltas from window features (idle vs interactive concat)."""
    freeze = json.loads(freeze_anchor_path().read_text(encoding="utf-8"))
    apps = freeze.get("apps") or {}
    evidence_root = Path(app_config.OUTPUT_DIR) / "evidence" / "dynamic"

    out: list[dict[str, Any]] = []
    for pkg, ent in sorted(apps.items()):
        if not isinstance(ent, dict):
            continue
        baseline_ids = [str(x) for x in (ent.get("baseline_run_ids") or [])]
        inter_ids = [str(x) for x in (ent.get("interactive_run_ids") or [])]
        if len(baseline_ids) < 1 or len(inter_ids) < 2:
            continue

        baseline_id = baseline_ids[0]
        interactive_ids = inter_ids[:2]  # frozen

        def load_series(rid: str) -> tuple[list[float], list[float], list[float]]:
            run_dir = evidence_root / rid
            inputs = load_run_inputs(run_dir)
            if not inputs or not inputs.pcap_path or not inputs.pcap_path.exists():
                return [], [], []
            duration_s = get_sampling_duration_seconds(inputs)
            if duration_s is None or duration_s <= 0:
                return [], [], []
            spec = WindowSpec(window_size_s=config.WINDOW_SIZE_S, stride_s=config.WINDOW_STRIDE_S)
            packets = extract_packet_timeline(inputs.pcap_path)
            rows, _ = build_window_features(packets, duration_s=float(duration_s), spec=spec)
            denom = float(spec.window_size_s) if spec.window_size_s > 0 else 1.0
            bps = [float(r.get("byte_count") or 0.0) / denom for r in rows]
            pps = [float(r.get("packet_count") or 0.0) / denom for r in rows]
            aps: list[float] = []
            for r in rows:
                bc = float(r.get("byte_count") or 0.0)
                pc = float(r.get("packet_count") or 0.0)
                aps.append((bc / pc) if pc > 0 else 0.0)
            return bps, pps, aps

        bps_idle, pps_idle, aps_idle = load_series(baseline_id)
        bps_int: list[float] = []
        pps_int: list[float] = []
        aps_int: list[float] = []
        for rid in interactive_ids:
            b, p, a = load_series(rid)
            bps_int.extend(b)
            pps_int.extend(p)
            aps_int.extend(a)

        b50_i = _quantile(bps_idle, 0.50)
        b95_i = _quantile(bps_idle, 0.95)
        b50_x = _quantile(bps_int, 0.50)
        b95_x = _quantile(bps_int, 0.95)

        p50_i = _quantile(pps_idle, 0.50)
        p95_i = _quantile(pps_idle, 0.95)
        p50_x = _quantile(pps_int, 0.50)
        p95_x = _quantile(pps_int, 0.95)

        a50_i = _quantile(aps_idle, 0.50)
        a95_i = _quantile(aps_idle, 0.95)
        a50_x = _quantile(aps_int, 0.50)
        a95_x = _quantile(aps_int, 0.95)

        out.append(
            {
                "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
                "bytes_p50_delta": round(b50_x - b50_i, 1),
                "bytes_p95_delta": round(b95_x - b95_i, 1),
                "pps_p50_delta": round(p50_x - p50_i, 2),
                "pps_p95_delta": round(p95_x - p95_i, 2),
                "pkt_size_p50_delta": round(a50_x - a50_i, 1),
                "pkt_size_p95_delta": round(a95_x - a95_i, 1),
            }
        )

    cols = [
        ("app", "App"),
        ("bytes_p50_delta", "Bytes/s Δ p50"),
        ("bytes_p95_delta", "Bytes/s Δ p95"),
        ("pps_p50_delta", "PPS Δ p50"),
        ("pps_p95_delta", "PPS Δ p95"),
        ("pkt_size_p50_delta", "PktSz Δ p50"),
        ("pkt_size_p95_delta", "PktSz Δ p95"),
    ]
    csv_path = tables_dir / "table_4_signature_deltas.csv"
    xlsx_path = tables_dir / "table_4_signature_deltas.xlsx"
    tex_path = tables_dir / "table_4_signature_deltas.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_4", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment="Table 4: Behavioral signature deltas (idle vs interactive), window stats (p50/p95 deltas).",
    )
    return csv_path, xlsx_path, tex_path


def _compute_static_posture_scores() -> dict[str, tuple[float, list[str]]]:
    """Compute Static Posture Score (0-100) per frozen app from baseline static_dynamic_plan.json.

    Context-only for Paper #2; never used as ML features.
    """
    freeze = json.loads(freeze_anchor_path().read_text(encoding="utf-8"))
    apps = freeze.get("apps") or {}
    evidence_root = Path(app_config.OUTPUT_DIR) / "evidence" / "dynamic"
    raw: list[tuple[str, int, int, int, float, list[str]]] = []
    for pkg, ent in sorted(apps.items()):
        if not isinstance(ent, dict):
            continue
        bid = (ent.get("baseline_run_ids") or [None])[0]
        if not bid:
            continue
        plan_path = evidence_root / str(bid) / "inputs" / "static_dynamic_plan.json"
        if not plan_path.exists():
            raw.append((pkg, 0, 0, 0, 0.0, ["missing_plan"]))
            continue
        obj = json.loads(plan_path.read_text(encoding="utf-8"))
        notes: list[str] = []

        ec = obj.get("exported_components") if isinstance(obj.get("exported_components"), dict) else {}
        e = ec.get("total")
        if e is None:
            try:
                e = sum(len(ec.get(k) or []) for k in ("activities", "services", "receivers", "providers"))
            except Exception:
                e = 0
            notes.append("exported_missing")

        perms = obj.get("permissions") if isinstance(obj.get("permissions"), dict) else {}
        dang = perms.get("dangerous")
        if isinstance(dang, list):
            p = len(dang)
        else:
            p = 0
            notes.append("dangerous_perms_missing")

        rf = obj.get("risk_flags") if isinstance(obj.get("risk_flags"), dict) else {}
        c = 1 if rf.get("uses_cleartext_traffic") is True else 0

        # SDK indicators are context-only. For frozen Paper #2 plans this is typically missing.
        s = 0.0
        sdk = obj.get("sdk_indicators") if isinstance(obj.get("sdk_indicators"), dict) else None
        if isinstance(sdk, dict) and sdk.get("score") is not None:
            try:
                s = float(sdk.get("score") or 0.0)
            except Exception:
                s = 0.0
                notes.append("sdk_indicators_invalid")
        else:
            notes.append("sdk_indicators_missing")

        raw.append((pkg, int(e or 0), int(p or 0), int(c), float(s), notes))

    es = [r[1] for r in raw]
    ps = [r[2] for r in raw]
    e_min, e_max = (min(es), max(es)) if es else (0, 0)
    p_min, p_max = (min(ps), max(ps)) if ps else (0, 0)

    out: dict[str, tuple[float, list[str]]] = {}
    for pkg, e, p, c, s, notes in raw:
        e_n = (float(e - e_min) / float(e_max - e_min)) if e_max > e_min else 0.0
        p_n = (float(p - p_min) / float(p_max - p_min)) if p_max > p_min else 0.0
        c_n = float(c)
        s_n = float(s)
        score = 100.0 * (0.25 * e_n + 0.25 * p_n + 0.25 * c_n + 0.25 * s_n)
        out[pkg] = (float(score), notes)
    return out


def _load_interactive_rdi_iforest() -> dict[str, float]:
    rows = _read_csv_rows(dataset_tables_dir() / "anomaly_prevalence_per_app_phase.csv")
    out: dict[str, float] = {}
    for r in rows:
        if r.get("phase") == "interactive" and r.get("model") == config.MODEL_IFOREST:
            try:
                out[str(r.get("package_name"))] = float(r.get("flagged_pct") or 0.0)
            except Exception:
                continue
    return out


def _rank_tertiles_4_4_4(
    values_by_pkg: dict[str, float | None],
    *,
    ascending: bool,
) -> tuple[dict[str, str], list[str]]:
    """Assign exact 4/4/4 rank-based tertiles (n=12).

    Ties are broken deterministically by package_name to preserve equal bin sizes.
    Missing values are excluded from grading and returned separately.
    """
    missing = sorted([pkg for pkg, v in values_by_pkg.items() if v is None])
    present = [(pkg, float(v)) for pkg, v in values_by_pkg.items() if v is not None]
    present.sort(key=lambda t: (t[1], t[0]) if ascending else (-t[1], t[0]))

    n = len(present)
    grade: dict[str, str] = {}
    if n <= 0:
        return grade, missing

    # Paper #2 expects n=12. Keep a defensive fallback for partial data.
    if n == 12:
        cuts = (4, 8)
    else:
        k = max(1, n // 3)
        cuts = (k, 2 * k)

    for i, (pkg, _) in enumerate(present):
        if i < cuts[0]:
            grade[pkg] = "Low"
        elif i < cuts[1]:
            grade[pkg] = "Medium"
        else:
            grade[pkg] = "High"
    return grade, missing


def _write_table_5_masvs_coverage(
    tables_dir: Path, *, provenance: dict[str, str]
) -> tuple[Path, Path, Path, list[dict[str, str]]]:
    """Table 5: MASVS category mapping derived from Phase B static reports.

    This is a context-only artifact. It does not affect Phase E ML features or scoring.
    """

    def _parse_iso_z(s: str) -> datetime | None:
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None

    def _pick_static_report(*, base_apk_sha256: str, package_name: str, plan_generated_at: str | None) -> Path | None:
        reports_dir = Path(app_config.DATA_DIR) / "static_analysis" / "reports"
        if not reports_dir.exists():
            return None
        candidates: list[Path] = []
        # Common filename patterns include <sha>_static-batch-...-<package>.json and <sha>_YYYYMMDD.json.
        for p in reports_dir.glob(f"{base_apk_sha256}*.json"):
            if not p.is_file():
                continue
            # Prefer filenames that name the package explicitly to avoid sha collisions across variants.
            if package_name in p.name:
                candidates.append(p)
        if not candidates:
            candidates = [p for p in reports_dir.glob(f"{base_apk_sha256}*.json") if p.is_file()]
        if not candidates:
            return None

        plan_ts = _parse_iso_z(plan_generated_at) if plan_generated_at else None

        scored: list[tuple[int, float, Path]] = []
        for p in candidates:
            try:
                obj = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
            md = obj.get("metadata") if isinstance(obj.get("metadata"), dict) else {}
            pkg = str(md.get("package_name") or "").strip()
            if pkg and pkg != package_name:
                continue
            prof = str(md.get("run_profile") or "").strip().lower()
            scope = str(md.get("run_scope") or "").strip().lower()
            # Strong preference for full/app runs (paper-grade Phase B).
            quality = 0
            if prof == "full":
                quality += 2
            if scope == "app":
                quality += 1

            gen = _parse_iso_z(str(obj.get("generated_at") or ""))
            if plan_ts and gen:
                # Prefer the static report closest at/behind plan generation time.
                delta = (plan_ts - gen).total_seconds()
                # Behind => non-negative; ahead => penalize heavily.
                dist = abs(delta) if delta >= 0 else (abs(delta) + 1e9)
            elif gen:
                # Fall back to recency.
                dist = -gen.timestamp()
            else:
                dist = float("inf")

            scored.append((quality, dist, p))
        if not scored:
            return None
        scored.sort(key=lambda t: (-t[0], t[1], t[2].name))
        return scored[0][2]

    freeze = json.loads(freeze_anchor_path().read_text(encoding="utf-8"))
    apps = freeze.get("apps") or {}
    evidence_root = Path(app_config.OUTPUT_DIR) / "evidence" / "dynamic"

    cats = ("PLATFORM", "NETWORK", "PRIVACY", "STORAGE", "CRYPTO", "RESILIENCE", "OTHER")
    out: list[dict[str, Any]] = []
    masvs_inputs: list[dict[str, str]] = []

    for pkg, ent in sorted(apps.items()):
        if not isinstance(ent, dict):
            continue
        bid = (ent.get("baseline_run_ids") or [None])[0]
        if not bid:
            continue
        plan_path = evidence_root / str(bid) / "inputs" / "static_dynamic_plan.json"
        if not plan_path.exists():
            continue
        plan = json.loads(plan_path.read_text(encoding="utf-8"))
        base_sha = str((plan.get("run_identity") or {}).get("base_apk_sha256") or "").strip()
        if not base_sha:
            continue
        plan_ts = str(plan.get("generated_at") or "").strip() or None
        report_path = _pick_static_report(base_apk_sha256=base_sha, package_name=pkg, plan_generated_at=plan_ts)

        counts_by_cat = {c: 0 for c in cats}
        sev = {"P0": 0, "P1": 0, "P2": 0, "NOTE": 0}
        total = 0
        notes: list[str] = []

        if not report_path:
            notes.append("static_report_missing")
        else:
            try:
                report = json.loads(report_path.read_text(encoding="utf-8"))
                findings = report.get("findings") if isinstance(report.get("findings"), list) else []
                for f in findings:
                    if not isinstance(f, dict):
                        continue
                    total += 1
                    cat = str(f.get("category_masvs") or "OTHER").strip().upper()
                    if cat not in counts_by_cat:
                        cat = "OTHER"
                    counts_by_cat[cat] += 1
                    sg = str(f.get("severity_gate") or "").strip().upper()
                    if sg in sev:
                        sev[sg] += 1
                masvs_inputs.append(
                    {
                        "package_name": pkg,
                        "base_apk_sha256": base_sha,
                        "static_report_path": str(report_path),
                        "static_report_sha256": _sha256_stream(report_path),
                    }
                )
            except Exception:
                notes.append("static_report_invalid_json")

        row: dict[str, Any] = {
            "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
            "package_name": pkg,
            "base_apk_sha256_8": base_sha[:8],
            "finding_count_total": total,
            "finding_count_p0": sev["P0"],
            "finding_count_p1": sev["P1"],
            "finding_count_p2": sev["P2"],
            "finding_count_note": sev["NOTE"],
            "notes": ";".join(notes) if notes else "",
        }
        for c in cats:
            row[f"finding_count_{c.lower()}"] = counts_by_cat[c]
        out.append(row)

    cols = [
        ("app", "App"),
        ("package_name", "Package"),
        ("base_apk_sha256_8", "Base SHA"),
        ("finding_count_total", "Finding count"),
        ("finding_count_p0", "P0"),
        ("finding_count_p1", "P1"),
        ("finding_count_p2", "P2"),
        ("finding_count_note", "NOTE"),
        ("finding_count_platform", "PLATFORM"),
        ("finding_count_network", "NETWORK"),
        ("finding_count_privacy", "PRIVACY"),
        ("finding_count_storage", "STORAGE"),
        ("finding_count_crypto", "CRYPTO"),
        ("finding_count_resilience", "RESILIENCE"),
        ("finding_count_other", "OTHER"),
        ("notes", "Notes"),
    ]
    csv_path = tables_dir / "table_5_masvs_coverage.csv"
    xlsx_path = tables_dir / "table_5_masvs_coverage.xlsx"
    tex_path = tables_dir / "table_5_masvs_coverage.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_5", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment=(
            "Table 5: MASVS category mapping from Phase B static findings (finding counts). "
            "Severity (P0–P2/NOTE) is tool-defined; not MASVS control criticality. "
            "Context-only; not a compliance checklist."
        ),
    )
    return csv_path, xlsx_path, tex_path, masvs_inputs


def _write_table_6_static_posture_scores(
    tables_dir: Path, *, provenance: dict[str, str]
) -> tuple[Path, Path, Path]:
    """Table 6: Static posture components and score (context-only, Fig B4 auditable)."""
    freeze = json.loads(freeze_anchor_path().read_text(encoding="utf-8"))
    apps = freeze.get("apps") or {}
    evidence_root = Path(app_config.OUTPUT_DIR) / "evidence" / "dynamic"

    raw: list[dict[str, Any]] = []
    for pkg, ent in sorted(apps.items()):
        if not isinstance(ent, dict):
            continue
        bid = (ent.get("baseline_run_ids") or [None])[0]
        if not bid:
            continue
        plan_path = evidence_root / str(bid) / "inputs" / "static_dynamic_plan.json"
        notes: list[str] = []
        if not plan_path.exists():
            raw.append(
                {
                    "package_name": pkg,
                    "exported_raw": 0,
                    "dangerous_raw": 0,
                    "cleartext_flag": 0,
                    "sdk_score": 0.0,
                    "notes": "missing_plan",
                }
            )
            continue

        obj = json.loads(plan_path.read_text(encoding="utf-8"))

        ec = obj.get("exported_components") if isinstance(obj.get("exported_components"), dict) else {}
        e = ec.get("total")
        if e is None:
            try:
                e = sum(len(ec.get(k) or []) for k in ("activities", "services", "receivers", "providers"))
            except Exception:
                e = 0
            notes.append("exported_missing_total")

        perms = obj.get("permissions") if isinstance(obj.get("permissions"), dict) else {}
        dang = perms.get("dangerous")
        if isinstance(dang, list):
            p = len(dang)
        else:
            p = 0
            notes.append("dangerous_perms_missing")

        rf = obj.get("risk_flags") if isinstance(obj.get("risk_flags"), dict) else {}
        c = 1 if rf.get("uses_cleartext_traffic") is True else 0

        s = 0.0
        sdk = obj.get("sdk_indicators") if isinstance(obj.get("sdk_indicators"), dict) else None
        if isinstance(sdk, dict) and sdk.get("score") is not None:
            try:
                s = float(sdk.get("score") or 0.0)
            except Exception:
                s = 0.0
                notes.append("sdk_indicators_invalid")
        else:
            notes.append("sdk_indicators_missing")

        raw.append(
            {
                "package_name": pkg,
                "exported_raw": int(e or 0),
                "dangerous_raw": int(p or 0),
                "cleartext_flag": int(c),
                "sdk_score": float(s),
                "notes": ";".join(notes) if notes else "",
            }
        )

    es = [int(r.get("exported_raw") or 0) for r in raw]
    ps = [int(r.get("dangerous_raw") or 0) for r in raw]
    e_min, e_max = (min(es), max(es)) if es else (0, 0)
    p_min, p_max = (min(ps), max(ps)) if ps else (0, 0)

    out: list[dict[str, Any]] = []
    for r in raw:
        e = int(r.get("exported_raw") or 0)
        p = int(r.get("dangerous_raw") or 0)
        c = int(r.get("cleartext_flag") or 0)
        s = float(r.get("sdk_score") or 0.0)
        e_n = (float(e - e_min) / float(e_max - e_min)) if e_max > e_min else 0.0
        p_n = (float(p - p_min) / float(p_max - p_min)) if p_max > p_min else 0.0
        score = 100.0 * (0.25 * e_n + 0.25 * p_n + 0.25 * float(c) + 0.25 * float(s))
        pkg = str(r.get("package_name") or "").strip()
        out.append(
            {
                "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
                "package_name": pkg,
                "exported_raw": e,
                "dangerous_raw": p,
                "cleartext_flag": c,
                "sdk_score": round(s, 3),
                "exported_norm": round(e_n, 3),
                "dangerous_norm": round(p_n, 3),
                "static_posture_score": round(float(score), 2),
                "notes": str(r.get("notes") or ""),
            }
        )

    cols = [
        ("app", "App"),
        ("package_name", "Package"),
        ("exported_raw", "Exported (raw)"),
        ("dangerous_raw", "Dangerous perms (raw)"),
        ("cleartext_flag", "Cleartext flag"),
        ("sdk_score", "SDK score"),
        ("exported_norm", "E norm"),
        ("dangerous_norm", "P norm"),
        ("static_posture_score", "StaticPostureScore"),
        ("notes", "Notes"),
    ]
    csv_path = tables_dir / "table_6_static_posture_scores.csv"
    xlsx_path = tables_dir / "table_6_static_posture_scores.xlsx"
    tex_path = tables_dir / "table_6_static_posture_scores.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_6", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment="Table 6: Static posture components and score (context-only; Fig B4 inputs).",
    )
    return csv_path, xlsx_path, tex_path


def _write_table_7_exposure_deviation_summary(
    tables_dir: Path, *, provenance: dict[str, str]
) -> tuple[Path, Path, Path]:
    """Table 7: interpretive Exposure–Deviation summary (no fused scalar)."""
    # Exposure: recompute from baseline plans to bin on full precision (Table 6 is rounded for readability).
    posture = _compute_static_posture_scores()
    exposure_by_pkg: dict[str, float | None] = {}
    exposure_notes_by_pkg: dict[str, list[str]] = {}
    for pkg, (score, notes) in posture.items():
        exposure_notes_by_pkg[pkg] = list(notes)
        # Treat missing plan as missing for bin computation (still reported with note).
        exposure_by_pkg[pkg] = None if "missing_plan" in notes else float(score)

    # Deviation comes from the dataset-level prevalence table (full precision), interactive only.
    rdi_rows = _read_csv_rows(dataset_tables_dir() / "anomaly_prevalence_per_app_phase.csv")
    if_rdi_by_pkg: dict[str, float | None] = {pkg: None for pkg in exposure_by_pkg}
    oc_rdi_by_pkg: dict[str, float | None] = {pkg: None for pkg in exposure_by_pkg}
    training_mode_if: dict[str, str] = {}
    for r in rdi_rows:
        pkg = str(r.get("package_name") or "").strip()
        if pkg not in if_rdi_by_pkg:
            continue
        if str(r.get("phase") or "").strip().lower() != "interactive":
            continue
        model = str(r.get("model") or "").strip()
        try:
            v = float(r.get("flagged_pct") or 0.0)
        except Exception:
            v = None
        if model == config.MODEL_IFOREST:
            if_rdi_by_pkg[pkg] = v
            training_mode_if[pkg] = str(r.get("training_mode") or "baseline_only")
        elif model == config.MODEL_OCSVM:
            oc_rdi_by_pkg[pkg] = v

    exposure_grade, exposure_missing = _rank_tertiles_4_4_4(exposure_by_pkg, ascending=True)
    deviation_grade_if, deviation_missing = _rank_tertiles_4_4_4(if_rdi_by_pkg, ascending=True)

    out: list[dict[str, Any]] = []
    for pkg in sorted(exposure_by_pkg.keys()):
        exposure = exposure_by_pkg.get(pkg)
        rdi_if = if_rdi_by_pkg.get(pkg)
        rdi_oc = oc_rdi_by_pkg.get(pkg)
        e_grade = "" if pkg in exposure_missing else exposure_grade.get(pkg, "")
        d_grade = "" if pkg in deviation_missing else deviation_grade_if.get(pkg, "")
        quadrant = f"{e_grade} Exposure + {d_grade} Deviation" if e_grade and d_grade else ""
        training_mode = training_mode_if.get(pkg, "")
        notes = []
        if training_mode and training_mode != "baseline_only":
            notes.append("union_fallback")
        if pkg in exposure_missing:
            notes.append("missing_exposure")
        if pkg in deviation_missing:
            notes.append("missing_deviation")
        # Preserve source notes for auditability (e.g., sdk_indicators_missing).
        notes.extend(exposure_notes_by_pkg.get(pkg) or [])

        out.append(
            {
                "app": config.DISPLAY_NAME_BY_PACKAGE.get(pkg, pkg),
                "package_name": pkg,
                "static_exposure_score": round(exposure, 2) if exposure is not None else None,
                "exposure_grade": e_grade,
                "rdi_if_interactive": round(rdi_if, 3) if rdi_if is not None else None,
                "deviation_grade_if": d_grade,
                "quadrant_label_if": quadrant,
                "rdi_ocsvm_interactive": round(rdi_oc, 3) if rdi_oc is not None else None,
                "training_mode_if": training_mode,
                "notes": ";".join(notes),
            }
        )

    # Render: keep raw full-precision values in CSV; rounding happens in xlsx/tex renderers.
    cols = [
        ("app", "App"),
        ("package_name", "Package"),
        ("static_exposure_score", "Exposure (StaticPostureScore)"),
        ("exposure_grade", "Exposure Grade"),
        ("rdi_if_interactive", "Deviation (RDI IF, interactive)"),
        ("deviation_grade_if", "Deviation Grade (IF)"),
        ("quadrant_label_if", "Regime (IF)"),
        ("rdi_ocsvm_interactive", "RDI OC-SVM (interactive)"),
        ("training_mode_if", "Train (IF)"),
        ("notes", "Notes"),
    ]
    csv_path = tables_dir / "table_7_exposure_deviation_summary.csv"
    xlsx_path = tables_dir / "table_7_exposure_deviation_summary.xlsx"
    tex_path = tables_dir / "table_7_exposure_deviation_summary.tex"
    _write_csv_with_provenance(csv_path, [k for k, _ in cols], out, provenance=provenance)
    _write_xlsx(xlsx_path, sheet_name="table_7", columns=cols, rows=out, provenance=provenance)
    _write_tex_table(
        tex_path,
        columns=cols,
        rows=out,
        provenance=provenance,
        caption_comment=(
            "Table 7: Interpretive Exposure–Deviation Summary over the frozen 12-app dataset. "
            "Exposure Grade and Deviation Grade are rank-based tertile bins (4/4/4) computed on full-precision values "
            "with deterministic tie-breaking by package_name. "
            "Grades and quadrant labels are interpretive overlays (not system outputs) and do not represent measured security risk."
        ),
    )
    return csv_path, xlsx_path, tex_path
